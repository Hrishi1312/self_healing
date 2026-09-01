"""Convert a markdown pipe table into an .xlsx workbook.

Usage:
    py md_table_to_excel.py <input.md> [output.xlsx]
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Excel refuses cell text longer than this.
_MAX_CELL = 32767
_SEPARATOR = re.compile(r"^\|?[\s:|-]+\|?$")


def split_row(line: str) -> list[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    cells, current, escaped = [], [], False
    for ch in line:
        if escaped:
            current.append(ch if ch == "|" else "\\" + ch)
            escaped = False
        elif ch == "\\":
            escaped = True
        elif ch == "|":
            cells.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
    if escaped:
        current.append("\\")
    cells.append("".join(current).strip())
    return cells


def parse_table(text: str) -> list[list[str]]:
    rows = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|"):
            continue
        if _SEPARATOR.match(stripped) and set(stripped) <= set("|-: "):
            continue
        rows.append(split_row(stripped))
    return rows


def write_xlsx(rows: list[list[str]], out_path: Path) -> None:
    width = max(len(r) for r in rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    for row in rows:
        padded = row + [""] * (width - len(row))
        ws.append([c[:_MAX_CELL] for c in padded])

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for idx in range(1, width + 1):
        longest = max(
            (len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(12, longest // 4), 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    src = Path(sys.argv[1])
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_suffix(".xlsx")
    rows = parse_table(src.read_text(encoding="utf-8"))
    if not rows:
        print(f"No markdown table found in {src}")
        return 1
    write_xlsx(rows, dst)
    print(f"Wrote {dst} ({len(rows) - 1} data rows x {max(len(r) for r in rows)} columns)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
